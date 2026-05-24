import streamlit as st
from datetime import datetime, timedelta
import io
import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload
import pandas as pd

# 1. CONFIGURACIÓN INICIAL Y SEGURIDAD
scope = ["https://www.googleapis.com/auth/drive"]
creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scope)
drive_service = build('drive', 'v3', credentials=creds)

FILE_ID = "1VyI_Sq6y2lfKUr8r0odOzEsMNuou610H"

RANGOS = [
    "Subcomisionados", "Mayores", "Capitanes", "Teniente", 
    "Subteniente", "Sargento 1ro", "Sargento 2do", 
    "Cabo 1ro", "Cabo 2do", "Agente"
]

# Función para cargar los datos actuales para la visualización
def cargar_datos_tabla():
    try:
        # Usamos el servicio ya autenticado para descargar el archivo
        request = drive_service.files().get_media(fileId=FILE_ID)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while done is False:
            status, done = downloader.next_chunk()
        
        fh.seek(0)
        # 1. Leemos el excel limpio sin asumir encabezados
        df = pd.read_excel(fh, engine='openpyxl', header=None)
        
        # 2. BUSCADOR AUTOMÁTICO DE TÍTULOS
        fila_titulos_idx = None
        for idx, row in df.iterrows():
            if 'RANGO' in [str(val).strip().upper() for val in row.values]:
                fila_titulos_idx = idx
                break
        
        if fila_titulos_idx is not None:
            titulos_reales = df.iloc[fila_titulos_idx].tolist()
            titulos_reales = [
                'Rango / Grado' if str(t).strip().upper() == 'RANGO' 
                else 'Evaluados / Listos' if str(t).strip().upper() == 'REEVALUACION'
                else '% Avance' if str(t).strip().upper() == 'PORCENTAJE'
                else 'Pendientes' if str(t).strip().upper() == 'PENDIENTES'
                else str(t) for t in titulos_reales
            ]
            
            df_datos = df.iloc[fila_titulos_idx + 1:].copy()
            df_datos.columns = titulos_reales
            df = df_datos
        else:
            st.warning("No se detectó la fila de encabezados automáticos.")
            return None

        # 3. LIMPIEZA BÁSICA DE FILAS VACÍAS
        df = df[df.iloc[:, 0].notna()]
        df = df[~df.iloc[:, 0].astype(str).str.contains('None|none', case=False, na=False)]
        
        # 4. CORTE AUTOMÁTICO HASTA EL TOTAL
        df = df.reset_index(drop=True)
        posicion_total = df[df.iloc[:, 0].astype(str).str.strip().str.upper() == 'TOTAL'].index
        
        if not posicion_total.empty:
            idx_total = posicion_total[0]
            # Cortamos temporalmente para asegurar que no entre nada DEBAJO del Total
            df = df.iloc[:idx_total + 1]

        # 5. LA EXCEPCIÓN: Eliminar "Subcomisionados" SOLO si está pegado al TOTAL (al final)
        if len(df) >= 2:
            # Revisamos la fila que quedó justamente ARRIBA del "TOTAL" (la penúltima fila)
            penultima_fila_texto = str(df.iloc[-2, 0]).strip().upper()
            
            if 'SUBCOMISIONADOS' in penultima_fila_texto:
                # Reconstruimos la tabla saltándonos esa fila penúltima
                fila_total = df.iloc[[-1]] # Guardamos la fila del TOTAL
                df_cuerpo = df.iloc[:-2]   # Guardamos todo lo de arriba
                df = pd.concat([df_cuerpo, fila_total]).reset_index(drop=True)
        
        # 6. Formatear números decimales automáticamente
        for col in df.columns:
            try:
                df[col] = pd.to_numeric(df[col], errors='ignore')
                if df[col].dtype == 'float64' or df[col].dtype == 'float32':
                    df[col] = df[col].round(1)
            except:
                pass
                
        # 7. Estética final para celdas vacías
        df = df.fillna("-")
        
        return df
    except Exception as e:
        st.error(f"Error en la automatización del reporte: {e}")
        return None

# 2. INTERFAZ DE USUARIO (Streamlit)
st.title("📋 Control Diario de Evaluaciones (.XLSX)")
st.subheader("Formulario de Registro Nacional")

# Inicializar almacenamiento de datos en la sesión para evitar recargas molestas
if "datos_cargar" not in st.session_state:
    st.session_state.datos_cargar = {rango: 0 for rango in RANGOS}
if "ultima_fecha_procesada" not in st.session_state:
    st.session_state.ultima_fecha_procesada = None

fecha_seleccionada = st.date_input("Seleccione la fecha del reporte:", datetime.now())

# VALIDACIÓN DEL RANGO DE DÍAS (Hoy y hasta 2 días atrás)
fecha_actual = datetime.now().date()
limite_pasado = fecha_actual - timedelta(days=5)

if fecha_seleccionada > fecha_actual:
    st.error("❌ No se pueden registrar datos de fechas futuras.")
elif fecha_seleccionada < limite_pasado:
    st.error(f"🔒 Esta fecha está bloqueada. Solo se pueden modificar reportes desde el {limite_pasado.strftime('%d de %B')}.")
else:
    meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    fecha_formateada = f"{fecha_seleccionada.day} de {meses[fecha_seleccionada.month - 1]}"

    # BOTÓN PARA CONSULTAR DATOS EXISTENTES
    # Esto evita que el programa esté descargando el Excel a cada segundo mientras te mueves por la interfaz
    if st.button("🔍 Cargar/Verificar datos de esta fecha") or st.session_state.ultima_fecha_procesada != fecha_formateada:
        with st.spinner("Buscando si existen registros previos en el Excel..."):
            try:
                # Descargar archivo a la memoria
                request = drive_service.files().get_media(fileId=FILE_ID)
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()

                fh.seek(0)
                wb = openpyxl.load_workbook(fh, data_only=True) # data_only=True para leer valores, no las fórmulas escritas
                ws = wb.active

                # Buscar la columna de la fecha
                col_idx = None
                max_col = ws.max_column
                for col in range(6, max_col + 1):
                    celda_valor = str(ws.cell(row=3, column=col).value).strip().lower()
                    if celda_valor == fecha_formateada.strip().lower():
                        col_idx = col
                        break

                # Si la columna existe, extraer los números reales del Excel
                if col_idx:
                    for idx, rango in enumerate(RANGOS):
                        valor_celda = ws.cell(row=4 + idx, column=col_idx).value
                        if valor_celda is not None:
                            try:
                                st.session_state.datos_cargar[rango] = int(valor_celda)
                            except (ValueError, TypeError):
                                st.session_state.datos_cargar[rango] = 0
                        else:
                            st.session_state.datos_cargar[rango] = 0
                    
                    # 💡 ESTE ES EL TRUCO CLAVE: Guardamos la fecha actual en el estado y forzamos el rediseño
                    st.session_state.ultima_fecha_procesada = fecha_formateada
                    st.rerun()  # <--- Esto obliga a Streamlit a pintar los números recién cargados
                    
                else:
                    # Si no existe la columna, poner todo en 0 para un registro nuevo
                    for rango in RANGOS:
                        st.session_state.datos_cargar[rango] = 0
                    st.session_state.ultima_fecha_procesada = fecha_formateada
                    st.toast(f"ℹ️ No hay datos previos para el {fecha_formateada}. Registro limpio.", icon="📝")
                    st.rerun() # <--- También redibujamos si es una columna limpia

            except Exception as e:
                st.error(f"Error al leer datos previos: {e}")

# Mostrar notificación persistente si los datos ya corresponden a la fecha seleccionada
    if st.session_state.ultima_fecha_procesada == fecha_formateada:
        # Verificar si hay algún número mayor a 0 para saber si vino con datos o limpia
        if any(st.session_state.datos_cargar.values()):
            st.success(f"📬 Datos del {fecha_formateada} cargados correctamente del Excel.")
        else:
            st.info(f"📝 Mostrando formulario limpio para el {fecha_formateada}.")

    # MOSTRAR EL FORMULARIO CON LOS VALORES CARGADOS
    st.write(f"### Evaluaciones para el {fecha_formateada}:")
    datos_nuevos = {}
    
    col1, col2 = st.columns(2)
    for i, rango in enumerate(RANGOS):
        with col1 if i < 5 else col2:
            # Aquí está el truco: el value por defecto ahora es lo que extrajimos del Excel (o 0 si es nuevo)
            datos_nuevos[rango] = st.number_input(
                f"{rango}:", 
                min_value=0, 
                step=1, 
                value=st.session_state.datos_cargar[rango],
                key=f"input_{rango}_{fecha_formateada}" # Clave única por fecha para evitar conflictos de Streamlit
            )

    if st.button("🚀 Guardar / Actualizar Reporte en Drive"):
        with st.spinner("Conectando a Drive y actualizando archivo..."):
            try:
                # 1. DESCARGAR EL EXCEL ACTUAL
                request = drive_service.files().get_media(fileId=FILE_ID)
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()

                fh.seek(0)
                wb = openpyxl.load_workbook(fh) # Aquí NO usamos data_only=True para conservar las fórmulas vivas del archivo
                ws = wb.active

                # 2. BUSCAR O CREAR COLUMNA
                col_idx = None
                max_col = ws.max_column
                for col in range(6, max_col + 2):
                    celda_fecha = ws.cell(row=3, column=col).value
                    if celda_fecha == fecha_formateada:
                        col_idx = col
                        break
                    elif celda_fecha is None:
                        col_idx = col
                        ws.cell(row=3, column=col, value=fecha_formateada)
                        break

                # 3. ESCRIBIR LOS VALORES (Los modificados y los que se quedaron igual)
                for idx, rango in enumerate(RANGOS):
                    fila_destino = 4 + idx
                    ws.cell(row=fila_destino, column=col_idx, value=datos_nuevos[rango])

                # 4. INYECTAR FÓRMULAS
                letra_col = openpyxl.utils.get_column_letter(col_idx)
                ws.cell(row=14, column=col_idx, value=f"=SUM({letra_col}4:{letra_col}13)")

                for idx in range(len(RANGOS)):
                    f = 4 + idx
                    ws.cell(row=f, column=3, value=f"=SUM(F{f}:{letra_col}{f})")
                    ws.cell(row=f, column=4, value=f"=(C{f}/B{f})*100")
                    ws.cell(row=f, column=5, value=f"=B{f}-C{f}")

                # 5. SUBIR A DRIVE
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                media = MediaIoBaseUpload(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', resumable=True)
                drive_service.files().update(fileId=FILE_ID, media_body=media).execute()

                # Actualizar la memoria local para que coincida con lo recién guardado
                for rango in RANGOS:
                    st.session_state.datos_cargar[rango] = datos_nuevos[rango]

                st.success(f"¡Reporte del {fecha_formateada} actualizado con éxito!")

            except Exception as e:
                st.error(f"Hubo un problema al guardar: {e}")

# --- SECCIÓN DE VISUALIZACIÓN (Al final del archivo) ---
st.divider()
st.subheader("📊 Monitoreo de Registros en Tiempo Real")

# Configuración para forzar 2 decimales y sufijo % en la columna % Avance
config_columnas = {
    "% Avance": st.column_config.NumberColumn(
        format="%.2f%%"
    )
}

# Función de estilo para resaltar el 100% en verde (éxito)
def resaltar_completado(val):
    return 'background-color: #c8e6c9;' if isinstance(val, (int, float)) and val >= 100 else ''

# Función de estilo para resaltar pendientes en rojo suave
def resaltar_pendientes(val):
    return 'background-color: #ffcdd2;' if isinstance(val, (int, float)) and val > 0 else ''

# Función para aplicar negrita a toda la fila si es un total
def estilo_negrita_totales(row):
    # Si la celda de la primera columna contiene 'TOTAL', aplicamos negrita a toda la fila
    es_total = 'TOTAL' in str(row['Rango / Grado']).upper()
    return ['font-weight: bold' if es_total else '' for _ in row]

# Botón para refrescar la tabla manualmente
if st.button("🔄 Actualizar Tabla"):
    df_actual = cargar_datos_tabla()
    if df_actual is not None:
        # Aplicamos el estilo verificando que las columnas existan para evitar KeyError
        styler = df_actual.style
        if '% Avance' in df_actual.columns:
            styler = styler.map(resaltar_completado, subset=['% Avance'])
        if 'Pendientes' in df_actual.columns:
            styler = styler.map(resaltar_pendientes, subset=['Pendientes'])
        
        df_estilizado = styler.apply(estilo_negrita_totales, axis=1)
        
        st.dataframe(df_estilizado, use_container_width=True, hide_index=True, 
                     column_config=config_columnas)
    else:
        st.warning("No se pudo cargar la tabla. Asegúrate de que existan datos.")

# Mostrar la tabla por defecto
df_vista = cargar_datos_tabla()
if df_vista is not None:
    st.write("📋 **Vista consolidada del estado de avance:**")
    # Aplicamos el estilo condicional para la vista por defecto
    styler_vista = df_vista.style
    if '% Avance' in df_vista.columns:
        styler_vista = styler_vista.map(resaltar_completado, subset=['% Avance'])
    if 'Pendientes' in df_vista.columns:
        styler_vista = styler_vista.map(resaltar_pendientes, subset=['Pendientes'])
    
    df_estilizado_vista = styler_vista.apply(estilo_negrita_totales, axis=1)

    st.dataframe(
        df_estilizado_vista, 
        use_container_width=True, 
        hide_index=True,
        column_config=config_columnas
    )

    # Gráfico de barras para visualizar el avance por rango
    st.divider()
    st.write("### 📈 Análisis Visual de Avance por Rango")
    st.bar_chart(df_vista, x="Rango / Grado", y="% Avance", color="#2e7d32")